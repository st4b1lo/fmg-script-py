import paramiko
from openpyxl import Workbook
import pandas as pd
import json
import requests
from requests.packages.urllib3.exceptions import InsecureRequestWarning
from tqdm import tqdm
import openpyxl
from getpass import getpass

# disable https warnings
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)


ssh = paramiko.SSHClient()
ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
ssh.connect(hostname=host, username=user, password=secret, port=port)

ssh = paramiko.SSHClient()
ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
ssh.connect(hostname=host, username=user, password=secret, port=port)

# execute device list
command = 'diagnose dvm device list'
stdin, stdout, stderr = ssh.exec_command(command, get_pty=True)
output = stdout.read().decode('utf-8')

lines = output.strip().split("\n")
filtered_lines = [line.strip() for line in lines if line.startswith('fmgfaz-managed') and line.strip()]

# creation excel file
workbook = Workbook()
sheet = workbook.active

for line in tqdm(filtered_lines, desc="Caricamento dati"):
    values = line.split()

    # Format the desired value as "adom/ADOM/device/NAME"
    adom = values[6]  # ADOM value
    device_name = values[5]  # NAME value
    formatted_value = f'adom/{adom}/device/{device_name}'

    # Append the formatted value to the sheet
    sheet.append([formatted_value])

# combine columns I, J, K e L in column I
for row in tqdm(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4), desc="Unione colonne"):
    combined_value = ""
    for cell in row:
        if cell.value:
            combined_value += str(cell.value) + " "
    combined_value = combined_value.strip()
    row[0].value = combined_value

# erase column B, C e D
sheet.delete_cols(2, 3)

output_file = "targets.xlsx"
workbook.save(output_file)

print(f"Output salvato nel file '{output_file}'.")

# Continue with the rest of the script

url = 'https://' + host + '/jsonrpc'

# load target from an excel file
try:
    targets_df = pd.read_excel('targets.xlsx', header=None)
except Exception as e:
    print("Si è verificato un errore durante la lettura del file Excel:", e)
    exit(1)

# extract target from column "A"
targets = targets_df.iloc[:, 0].tolist()

payload = {
    "id": 1,
    "method": "exec",
    "params": [
        {
            "data": {
                "passwd": secret,
                "user": user
            },
            "url": "/sys/login/user"
        }
    ],
    "session": "string"
}

response = requests.post(url, json=payload, verify=False)

if response.status_code == 200:
    data = response.json()
    session_id = data.get('session')
    if session_id is None:
        print("Errore: Chiave 'session' non presente nel JSON restituito.")
        exit(1)
else:
    print('Errore nella richiesta:', response.status_code)
    exit(1)

new_payload = {
    "id": 2,
    "method": "exec",
    "params": [
        {
            "data": {
                "action": "get",
                "resource": "/api/v2/cmdb/vpn.ssl/settings",
                "target": targets
            },
            "url": "sys/proxy/json"
        }
    ],
    "session": session_id
}

response = requests.post(url, json=new_payload, verify=False)

if response.status_code == 200:
    data = response.json()
    # Save output as JSON
    with open('new_output.json', 'w') as json_file:
        json.dump(data, json_file)

    # Convert JSON data to DataFrame
    df = pd.json_normalize(data['result'][0]['data'], errors='ignore')

    # Select specific columns
    df = df[['response.version', 'response.serial', 'response.results.status',
             'response.results.port', 'response.results.tunnel-ip-pools',
             'response.results.source-interface', 'target']]

    # Save DataFrame as Excel file
    json_sslvpn_file = 'json_sslvpn.xlsx'
    df.to_excel(json_sslvpn_file, index=False)
    print(f"Output salvato nel file '{json_sslvpn_file}'.")

    # Load the json_sslvpn.xlsx file
    workbook_sslvpn = openpyxl.load_workbook(json_sslvpn_file)
    sheet_sslvpn = workbook_sslvpn.active

    # Load the targets.xlsx file
    workbook_targets = openpyxl.load_workbook('targets.xlsx')
    sheet_targets = workbook_targets.active

    # Find the column index to insert the ADOM column
    column_index = sheet_sslvpn.max_column + 1

    # Get the values from the ADOM column in targets.xlsx
    adom_values = [row[0].value for row in sheet_targets.iter_rows(min_row=1, max_row=sheet_targets.max_row)]

    # Insert the ADOM values in the ADOM column of json_sslvpn.xlsx
    for index, adom_value in enumerate(adom_values, start=2):
        sheet_sslvpn.cell(row=index, column=column_index, value=adom_value)

    # Save the modified json_sslvpn.xlsx file
    workbook_sslvpn.save(json_sslvpn_file)
    print(f"The 'ADOM' column has been added to '{json_sslvpn_file}' and populated with data from 'targets.xlsx'.")

else:
    print('Errore nella richiesta:', response.status_code)

logout_payload = {
    "id": 3,
    "method": "exec",
    "params": [
        {
            "data": {
                "url": "/sys/logout"
            },
            "url": "/sys/logout"
        }
    ],
    "session": session_id
}

response = requests.post(url, json=logout_payload, verify=False)

if response.status_code == 200:
    data = response.json()
else:
    print('Errore nella richiesta:', response.status_code)

print("Script execution completed.")
